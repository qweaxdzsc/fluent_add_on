
# TXT to XLS
def process_data(txt_name, path):
    import numpy as np

    def txt_in(filename):
        f = open(filename, encoding='utf-8')    # open
        txt = f.readlines()                     # read

        row_number = len(txt)
        for i in range(row_number):             # process txt
            txt[i] = txt[i].strip()             # strip blanks before and after
            txt[i] = txt[i].split()             # split line into list

        return txt

    txt = txt_in(txt_name)

    # Partitioning txt function
    def partitioning(txt, part_name):
        exist_factor = 0
        divide_sign = []

        for line in txt:
            if part_name in line:               # find part
                title_index = txt.index(line)
                title_line = txt[title_index]   # find line
                exist_factor += 1               # exist = 1
                break
        if exist_factor == 1:                   # if can be found
            txt_down = txt[title_index:]        # from the line have part_name
            for line in txt_down:
                if '--' in line[0]:             # use part_name and "--" to isolate part
                    divide_sign.append(txt_down.index(line))          # get index of top and bottom line
                if len(divide_sign) > 1:                              # find bottom, break
                    break
                    # print(divide_sign)
            txt_taget = txt_down[divide_sign[0]+1:divide_sign[1]]     # isolate part
            txt_taget = np.array(txt_taget)                           # use np array to manipulate

            txt_taget_string = txt_taget[:, 0]
            txt_taget_string = list(txt_taget_string)

            txt_taget_data = txt_taget[:, 1]
            txt_taget_data = list(map(float, txt_taget_data))         # float the data
        else:
            title_line, txt_taget_string, txt_taget_data = ['not-exist'], ['not-exist'], ['not-exist']

        return title_line, txt_taget_string, txt_taget_data, exist_factor

    # Partitioning volume flow rate

    title_line, volume_string, volume_data, volume_exist = partitioning(txt, 'Volumetric')

    # volume data process
    def unit_transform(data):
        data = np.array(data)
        data_liter_per_second = data*1000

        return data_liter_per_second

    def total_volume(data):
        """volume data, make it all nagetive value equal to zero"""
        total_volume = sum(np.maximum(data, 0))
        return total_volume

    # decimal function
    def decimal(raw_data, number):
        decimaled_list = [0 for i in raw_data]
        if number == 1:
            for i in range(len(raw_data)):
                decimaled_list[i] = '%.1f' % raw_data[i]
        elif number == 2:
            for i in range(len(raw_data)):
                decimaled_list[i] = '%.2f' % raw_data[i]
        elif number == 3:
            for i in range(len(raw_data)):
                decimaled_list[i] = '%.3f' % raw_data[i]

        return decimaled_list

    def add_symbol(string_ratio):
        percentaging_data = string_ratio + '%'         # add'%' to percentage value

        return percentaging_data

    def percentaging(raw_data, total_volume):          # whole percentaging process
        raw_percent = abs(raw_data/total_volume)*100

        def str_1(raw_data):
            str_1 = '%.1f'%raw_data

            return str_1
        string_ratio = list(map(str_1, raw_percent))

        processed_data = list(map(add_symbol, string_ratio))

        return processed_data

    if volume_exist == 1:                                                 # confirm volume part existence
        Liter_per_second = unit_transform(volume_data)                    # get L/S
        total_volume = total_volume(Liter_per_second)                     # get total volume

        Liter_per_second = list(Liter_per_second)
        Liter_per_second.append(total_volume)
        Liter_per_second_de = decimal(Liter_per_second, 1)                # add total to Liter_ array

        percentage_array = percentaging(Liter_per_second, total_volume)   # get percentage_array
        volume_string = list(volume_string)
        volume_string.append('Total')                                     # get volume part name

    else:
        # if not exist
        volume_string, Liter_per_second_de, percentage_array = ['not-exist'], ['not-exist'], ['not-exist']

    # Partitioning static pressure
    sp_title, sp_string, sp_data, sp_exist = partitioning(txt, 'Static')

    # static pressure data process
    if sp_exist == 1:
        sp_data_de = decimal(sp_data, 1)
    else:
        sp_data_de = ['not-exist']

    # Partitioning total pressure
    tp_title, tp_string, tp_data, tp_exist = partitioning(txt, 'Total')

    if tp_exist == 1:
        tp_data_de = decimal(tp_data, 1)
    else:
        tp_data_de = ['not-exist']

    # Partitioning uniformity
    uni_title, uni_string, uni_data, uni_exist = partitioning(txt, 'Uniformity')

    if uni_exist == 1:
        uni_data_de = decimal(uni_data, 3)
    else:
        uni_data_de = ['not-exist']

    # Partitioning Fan Moment
    def find_moment(txt):
        exist_factor = 0
        for line in txt:                                  # find moment
            if 'Moment' and 'Axis' in line:               # get that line
                part_index = txt.index(line)
                exist_factor += 1
                break
        if exist_factor == 1:                             # check existence
            moment_raw = txt[part_index+3][3]             # locate torque value
            try:
                moment = '%.3f'% abs(float(moment_raw))   # in case wrong location
            except Exception as e:
                error = "Wrong, Error:%s"%e
                return error
            else:
                return moment
        else:
            return 'not-exist'

    fan_moment = [find_moment(txt)]                       # fan moment

    # save all string and data into list(matrix)
    global matrix
    matrix = volume_string, Liter_per_second_de, percentage_array, sp_string, sp_data_de, tp_string, tp_data_de, uni_string, uni_data_de, fan_moment
    np.save(path + 'data_matrix', matrix, allow_pickle=True)

    return matrix


def get_xls(path, sheet_name, excel_name, data_name):
    # write to excel
    import numpy as np
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill
    import copy


    # create new excel

    def create_excel(excel_name, sheet_name):

        workbook = openpyxl.Workbook()                               # 创建一个Workbook对象，相当于创建了一个Excel文件

        worksheet = workbook.create_sheet(title='sheet2', index=0)   # create sheet in first
        worksheet.title = sheet_name

        workbook.save(filename=excel_name)                           # save to path
        print('create new excel:', excel_name, 'create new sheet:', sheet_name)


    # load excel

    def load_excel(excel_name, sheet_name):

        workbook = openpyxl.load_workbook(excel_name)                   # load excel in
        print('all sheet name:', workbook.sheetnames)                   # get all sheet's names

        worksheet = workbook[sheet_name]                                # use sheet name to find sheet
        print('opened excel:', excel_name, 'with sheet:', sheet_name)

        return workbook, worksheet


    # input data

    def write_excel(matrix, worksheet, head_line, shift_right, shift_down, data_title):

        exist_row_index = worksheet.max_row                     # get exist table's max row index
        print('current max row_index:', exist_row_index)
        shift_down = exist_row_index + shift_down               # shift new table under the exist table

        # data title
        worksheet.merge_cells(start_row=1+shift_down, start_column=1, end_row=1+shift_down+7, end_column=3)   # merge cell
        worksheet.cell(1+shift_down,  1,  data_title)           # input data title

        title_index = 'A' + str(1+shift_down)
        worksheet2[title_index].font = Font(size=16, italic=True, bold=True)        # change title font
        # head line
        for j in range(len(head_line)):
            worksheet.cell(1+shift_down, j+1+shift_right, head_line[j])             # input head line
            head_index = get_column_letter(j+1+shift_right) + str(1+shift_down)
            worksheet2[head_index].fill = PatternFill("solid", fgColor="D1EEEE")    # change head line color
            worksheet2[head_index].font = Font(size=12, bold=True)                  # change head line font
        # matrix to excel
        for j in range(len(matrix)):
            for i in range(len(matrix[j])):
                worksheet.cell(i+2+shift_down,  j+1+shift_right, str(matrix[j][i])) # input all data into cell

    # Modify excel style
    def auto_width(matrix, worksheet, head_line, shift_right):                      # auto adjust column width
        # get width of each cell into width_matrix
        width = copy.deepcopy(matrix)
        for i in range(len(matrix)):
            for j in range(len(matrix[i])):
                width[i][j] = len(matrix[i][j])                  # all cell width matrix

        col_width = np.zeros(len(matrix))                        # create column width array have same length with matrix
        head_width = np.zeros(len(head_line))                    # head_line width matrix

        for i in range(len(width)):
            col_letter = get_column_letter(i+1+shift_right)     # get column sequence letter
            col_width[i] = max(width[i])                        # save max width of each column into col_width
            head_width[i] = len(head_line[i])                   # get heading all column width
            # modify worksheet's each column with max width
            worksheet.column_dimensions[col_letter].width = max(col_width[i], head_width[i])*1.2 + 2


    # begin writing excel
    Excel_output_path = path + "%s.xlsx" % excel_name
    create_excel(Excel_output_path, sheet_name)                                     # create excel
    excel, worksheet2 = load_excel(Excel_output_path, sheet_name)                   # load sheet into worksheet

    head_line = ['', 'Volume(L/S)', 'Percentage(%)', 'Static Pressure', '(Pa)',
                 'Total Pressure', '(Pa)', 'Uniformity', '', 'Torque(N/m)']         # define head_line
    shift_right = 3
    shift_down = 5
    write_excel(matrix, worksheet2, head_line, shift_right, shift_down, data_name)  # write sheet
    auto_width(matrix, worksheet2, head_line, shift_right)                          # adjust column width

    excel.save(filename=Excel_output_path)    # save
