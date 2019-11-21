# import timeit
# import numpy as np
#
# # 待测试的函数
# def add():
#     import openpyxl
#     wb = openpyxl.Workbook()
#     worksheet = wb['Sheet']
#
#     worksheet.cell(1, 1, 2)
#
#     wb.save(filename=r'C:\Users\BZMBN4\Desktop\test.xlsx')
#
# # stmt 需要测试的函数或语句，字符串形式
# # setup 运行的环境，本例子中表示 if __name__ == '__main__':
# # number 被测试的函数或语句，执行的次数，本例表示执行100000次add()。省缺则默认是10000次
# # repeat 测试做100次
# # 综上：此函数表示 测试 在if __name__ == '__main__'的条件下，执行100000次add()消耗的时间，并把这个测试做100次,并求出平均值
#
# t = timeit.repeat(stmt="add()", setup="from __main__ import add", number=1, repeat=1)
#
# print(t)
# print(sum(t) / len(t))

from datetime import date
import os

from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis

wb = Workbook()
ws = wb.active

rows = [
    ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
    [date(2015,9, 1), 40, 30, 25],
    [date(2015,9, 2), 40, 25, 30],
    [date(2015,9, 3), 50, 30, 45],
    [date(2015,9, 4), 30, 25, 40],
    [date(2015,9, 5), 25, 35, 30],
    [date(2015,9, 6), 20, 40, 35],
]

for row in rows:
    ws.append(row)


c2 = LineChart()
c2.title = "Date Axis"
c2.style = 12
c2.y_axis.title = "Size"
c2.y_axis.crossAx = 500
c2.x_axis = DateAxis(crossAx=100)
c2.x_axis.number_format = 'd-mmm'
c2.x_axis.majorTimeUnit = "days"
c2.x_axis.title = "Date"

data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=7)
c2.add_data(data, titles_from_data=True)

dates = Reference(ws, min_col=1, min_row=2, max_row=7)
c2.set_categories(dates)

ws.add_chart(c2, "A20")


path = r"C:/Users/BZMBN4/Desktop/line.xlsx"
wb.save(path)
os.system(path)


