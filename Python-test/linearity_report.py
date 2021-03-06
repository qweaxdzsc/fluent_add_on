import time

from numpy import linspace
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    LineChart,
    ScatterChart,
    Reference,
    Series,
)
import os


class LinReport(object):
    def __init__(self, project_address, project_name, version_name, start_percent=10, end_percent=90, process_point=9):
        self.project_address = project_address
        self.project_name = project_name
        self.version_name = version_name
        self.start_percent = start_percent
        self.end_percent = end_percent
        self.process_point = process_point
        self.whole_name = project_name + version_name

        self.get_report()

    def get_report(self):
        self.temp_dict = {}                 # create empty dict
        self.form_angle_array()             # create valve open angle array
        for i in self.angle_array:          # extract data from all result file
            result_file = r'{project_address}\lin_case\{project_name}_{version_name}_{angle}\result\{project_name}_{angle}.txt' \
                .format(project_address=self.project_address, project_name=self.project_name,
                        version_name=self.version_name, angle=i)

            temp_angle = self.txt_to_dict(result_file)
            self.temp_dict[i] = temp_angle          # form one temperature dict

        self.create_excel()
        self.write_sheet()
        self.auto_width_sheet()                                                 # auto adjust width for excel
        # self.add_line_chart(whole_name + '_linearity')                          # create line chart for temp data
        self.add_scatter_chart(whole_name + '_linearity')
        self.save_open_excel(project_address + '\\' + whole_name + '.xlsx')

    def form_angle_array(self):
        angle = linspace(self.start_percent, self.end_percent, self.process_point)
        self.angle_array = [int(i) for i in angle]

    def txt_to_dict(self, txt_path):
        with open(txt_path, 'r') as f:
            lines = f.readlines()
            line_number = 0

            while 'Static Temperature' not in lines[line_number]:   # find temperature area and first temperature row
                lines[line_number] = lines[line_number].strip()
                line_number += 1
            line_number += 2
            first_row_number = line_number

            while '-----' not in lines[line_number]:                # last temperature row
                lines[line_number] = lines[line_number].strip()
                if not lines[line_number]:
                    break
                line_number += 1
            last_row_number = line_number

            temp_angle = {}
            for i in lines[first_row_number:last_row_number]:       # form dict
                i = i.split()
                temp_angle[i[0]] = i[1]

            return temp_angle

    def create_excel(self):
        self.wb = Workbook()
        self.sheet = self.wb['Sheet']

    def write_sheet(self):
        header = ['open percentage']
        outlet_list = list(self.temp_dict[self.angle_array[0]].keys())
        header.extend(outlet_list)

        for i in range(len(header)):
            self.sheet.cell(1, i+1, header[i]+'(°C)')

        second_row = [0 for i in outlet_list]
        second_row.insert(0, 0)
        for i in range(len(second_row)):
            self.sheet.cell(2, i + 1, second_row[i])

        for i in range(len(self.angle_array)):
            self.sheet.cell(i+3, 1, self.angle_array[i]/100)

        for i in range(len(self.angle_array)):
            for j in range(len(outlet_list)):
                temperature_k = float(self.temp_dict[self.angle_array[i]][outlet_list[j]])
                temp_c = round(temperature_k, 1)-273
                self.sheet.cell(3+i, 2+j, temp_c)

        last_row = [75 for i in outlet_list]
        last_row.insert(0, 1)
        for i in range(len(last_row)):
            self.sheet.cell(len(self.angle_array)+3, i + 1, last_row[i])

        for i in range(2, self.sheet.max_row + 1):
            self.sheet.cell(i, 1).number_format = '0%'
            # self.sheet.cell(i, 1).number_format = numbers.FORMAT_PERCENTAGE

    def auto_width_sheet(self):
        col_number = self.sheet.max_column
        for i in range(col_number):
            col_letter = get_column_letter(i+1)
            self.sheet.column_dimensions[col_letter].width = 16

    def save_open_excel(self, excel_path):
        self.wb.save(filename=excel_path)
        # os.system(excel_path)

    def add_line_chart(self, title, min_col=2, min_row=1):
        chart = LineChart()
        chart.height = 10
        chart.width = 18
        chart.title = title
        chart.style = 10                                                             # 线条的style
        chart.y_axis.title = 'Temperature'
        chart.x_axis.title = "Open percentage"

        data = Reference(self.sheet, min_col=min_col, min_row=min_row,
                         max_col=self.sheet.max_column, max_row=self.sheet.max_row)  # 图像的数据 起始行、起始列、终止行、终止列
        chart.add_data(data, titles_from_data=True)
        angles = Reference(self.sheet, min_col=1, min_row=2, max_row=self.sheet.max_row)
        chart.set_categories(angles)
        self.sheet.add_chart(chart, "A15")

    def add_scatter_chart(self, title, min_col=1, min_row=1):
        chart = ScatterChart()
        chart.title = title
        chart.height = 10
        chart.width = 18
        chart.style = 2  # 线条的style（8种颜色循环，1：灰度，2：异色，3-8：同色渐变-蓝，棕红，绿，紫，青，橙；1-8线条较细，9-16加粗）
        chart.y_axis.title = 'Temperature'
        chart.x_axis.title = "Open percentage"
        chart.x_axis.scaling.max = 1

        xvalues = Reference(self.sheet, min_col=min_col, min_row=min_row + 1, max_row=self.sheet.max_row)
        for i in range(2, self.sheet.max_column + 1):  # 数据列循环
            yvalues = Reference(self.sheet, min_col=i, min_row=min_row, max_row=self.sheet.max_row)
            series = Series(yvalues, xvalues, title_from_data=True)
            chart.series.append(series)

        self.sheet.add_chart(chart, "A16")  # 将图表添加到 sheet中


if __name__ == "__main__":
    start_time = time.time()
    project_address = r"G:\_HAVC_Project\D2U-2\D2U-2_lin_vent\D2U-2_V47_lin_vent"
    project_name = 'D2U-2'
    version_name = 'V47_lin_vent'

    whole_name = project_name + '-' + version_name
    Linearity_report = LinReport(project_address, project_name, version_name, 10, 60, 6)
    end_time = time.time()
    print('use_time:', end_time - start_time)