import os
import openpyxl
import win32com.client as client


class CFDProject(object):
    def __init__(self, project_name, create_path, mode_number=7):
        self.project_name = project_name
        self.create_path = create_path
        self.project_folder = '%s/%s' % (self.create_path, self.project_name)
        self.mode_number = mode_number
        self.modes = ['vent', 'foot', 'bi_level', 'defrost', 'defog', 'tri_level', 'hi_level']

    def create_project_folder(self):
        if os.path.exists(self.project_folder):
            print('project folder already exist:', self.project_folder)
        else:
            os.makedirs(self.project_folder)

    def create_sub_folder(self):
        for index, mode in enumerate(self.modes[:self.mode_number]):
            new_index = '%02d' % (index + 1)
            sub_folder = '%s/%s_%s_%s' % (self.project_folder, self.project_name, new_index, mode)
            new_lin_index = '%02d' % (index + 1 + self.mode_number)
            sub_lin_folder = '%s/%s_%s_lin_%s' % (self.project_folder, self.project_name, new_lin_index, mode)
            if not os.path.exists(sub_folder):
                os.makedirs(sub_folder)
            if not os.path.exists(sub_lin_folder):
                os.makedirs(sub_lin_folder)

    def create_record_excel(self):
        coefficient_book = openpyxl.Workbook()
        default_sheet = coefficient_book['Sheet']
        coefficient_book.remove(default_sheet)
        porous_sheet = coefficient_book.create_sheet('porous resistance')
        direction_sheet = coefficient_book.create_sheet('axis')
        duct_sheet = coefficient_book.create_sheet('duct_resistance')
        request_sheet = coefficient_book.create_sheet('request')

        direction_sheet.cell(1, 1, 'Version and mode')
        direction_sheet.cell(1, 2, 'fan')
        direction_sheet.cell(1, 3, 'evap')
        direction_sheet.cell(1, 4, 'hc')
        direction_sheet.cell(1, 5, 'filter')

        coefficient_book.save('%s/%s_coefficient.xlsx' % (self.project_folder, self.project_name))

        version_book = openpyxl.Workbook()
        default_sheet = version_book['Sheet']
        version_book.remove(default_sheet)
        for index, mode in enumerate(self.modes[:self.mode_number]):
            new_sheet = version_book.create_sheet(mode, index)
            new_lin_sheet = version_book.create_sheet('lin_%s' % mode, index + self.mode_number)

        version_book.save('%s/%s_version.xlsx' % (self.project_folder, self.project_name))

    def create_share_project(self, share_drive_path):
        share_project_folder = '%s/%s' % (share_drive_path, self.project_name)
        if os.path.exists(share_project_folder):
            print('project folder already exist:', share_project_folder)
        else:
            os.makedirs(share_project_folder)

        airflow_folder = '%s/01_Airflow' % share_project_folder
        lin_folder = '%s/02_Linearity' % share_project_folder

        for index, mode in enumerate(self.modes[:self.mode_number]):
            new_index = '%02d' % (index + 1)
            sub_folder = '%s/%s_%s_%s' % (airflow_folder, self.project_name, new_index, mode)
            sub_lin_folder = '%s/%s_%s_lin_%s' % (lin_folder, self.project_name, new_index, mode)
            if not os.path.exists(sub_folder):
                os.makedirs(sub_folder)
            if not os.path.exists(sub_lin_folder):
                os.makedirs(sub_lin_folder)

        other_folder = '%s/03_Other' % share_project_folder
        if not os.path.exists(other_folder):
            os.makedirs(other_folder)
        # ------------------create shortcut-------------------------------------------
        """filename should be abspath, or there will be some strange errors"""
        shell = client.Dispatch("WScript.Shell")
        version_xls = '%s/%s_version.xlsx' % (self.project_folder, self.project_name)
        version_xls_shortcut = '%s/%s_version.xlsx.lnk' % (share_project_folder, self.project_name)
        shortcut = shell.CreateShortCut(version_xls_shortcut)   # short cut path
        shortcut.TargetPath = os.path.abspath(version_xls)      # original file target
        shortcut.save()


if __name__ == "__main__":
    project_name = r"GX18"
    create_path = r"G:\_HAVC_Project"
    share_drive_path = r"S:\PE\Engineering database\CFD\02_Projects"
    new_project = CFDProject(project_name, create_path, mode_number=3)
    new_project.create_project_folder()
    new_project.create_sub_folder()
    new_project.create_record_excel()
    new_project.create_share_project(share_drive_path)
