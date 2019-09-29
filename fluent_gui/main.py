#!/usr/bin/python
# -*- coding: utf-8 -*-

import sys
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QBasicTimer
from PyQt5.QtGui import QTextCursor


from easy_test import Ui_MainWindow
from rename_tip import Ui_tip_widget
from K_cal import Ui_K_calculator
from unit_convertor import Ui_unit_converter
from porous_model import Ui_porous_model_form
import time
import os
import fluent_tui
import qdarkstyle


class MyMainWindow(QMainWindow, Ui_MainWindow):
    time_running = pyqtSignal(int)

    def __init__(self, parent=None):
        super(MyMainWindow, self).__init__(parent)
        self.setupUi(self)

        self.mode_ui()
        self.account_info()
        self.get_date()

        self.pamt = {}
        self.body_list = []
        self.btn()

    def mode_ui(self):
        self.mode_info_frame.show()
        self.start_btn.hide()
        self.mass_inlet.hide()
        self.return_btn.hide()
        self.evap_c.hide()
        self.hc_c.hide()
        self.solver_btn.hide()
        self.valve_c.hide()
        self.temp_c.hide()
        self.actionstop.setEnabled(False)

    def default_part_tree(self):
        self.part_tree.topLevelItem(0).setCheckState(0, QtCore.Qt.Unchecked)
        self.part_tree.topLevelItem(1).setCheckState(0, QtCore.Qt.Checked)
        self.part_tree.topLevelItem(2).setCheckState(0, QtCore.Qt.Unchecked)
        self.part_tree.topLevelItem(3).setCheckState(0, QtCore.Qt.Unchecked)
        self.part_tree.topLevelItem(4).setCheckState(0, QtCore.Qt.Unchecked)
        self.part_tree.topLevelItem(5).setCheckState(0, QtCore.Qt.Unchecked)
        self.part_tree.topLevelItem(6).setCheckState(0, QtCore.Qt.Unchecked)
        self.part_tree.topLevelItem(7).setCheckState(0, QtCore.Qt.Checked)
        self.part_tree.topLevelItem(8).setCheckState(0, QtCore.Qt.Checked)
        self.part_tree.topLevelItem(9).setCheckState(0, QtCore.Qt.Unchecked)
        self.part_tree.topLevelItem(10).setCheckState(0, QtCore.Qt.Unchecked)
        self.part_tree.topLevelItem(11).setCheckState(0, QtCore.Qt.Unchecked)
        self.distrib_number.setValue(1)
        self.outlet_number.setValue(1)

    def account_info(self):
        user = os.environ.get("USERNAME")
        self.username_label.setText('欢迎大佬%s' % (user))
        welcome_list = ['Hi,大佬%s，我们又见面了' % (user), 'Hello,欢迎大佬%s' % (user)]
        import random
        welcome_word = random.choice(welcome_list)
        self.interact_edit.append(welcome_word)

    def get_date(self):
        import time
        today = time.strftime('%y%m%d', time.localtime(time.time()))
        self.version_date_edit.setText(today)

    def btn(self):
        self.actionimport.triggered.connect(self.import_info)
        self.actionexport.triggered.connect(self.export_pamt)
        self.actionsolve.triggered.connect(self.quick_solve)
        self.actionstop.triggered.connect(self.force_stop)
        self.actiondarkstyle.triggered.connect(self.darkstyle)
        self.project_address_explore.clicked.connect(self.case_address)

        self.quick_distribfc_btn.toggled.connect(self.quick_distrib_judge)
        self.quick_distribfh_btn.toggled.connect(self.quick_distrib_judge)
        self.quick_distribbil_btn.toggled.connect(self.quick_distrib_judge)
        self.quick_distriblin_btn.toggled.connect(self.quick_distrib_judge)
        self.finish_mode_info_btn.clicked.connect(self.into_CAD)

        self.unit_btn.clicked.connect(self.unit_convert)
        self.start_btn.clicked.connect(self.begin)
        self.return_btn.clicked.connect(self.mode_ui)
        self.solver_btn.clicked.connect(self.solver)

        self.show_workflow_btn.clicked.connect(self.test)
        self.choose_evap_btn.clicked.connect(lambda: self.append_text('功能未开放,敬请期待'))
        self.actionalter_default_parameter.triggered.connect(lambda: self.append_text('功能未开放,敬请期待'))

    def test(self):
        self.porous_model = Ui_porous()
        self.porous_model.show()
        self.append_text('功能未开放,敬请期待')

        pass

    def darkstyle(self):
        app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
        self.actiondarkstyle.setDisabled(True)
        self.append_text('进入暗色主题')

    def append_text(self, msg):
        self.interact_edit.append(msg)
        self.interact_edit.moveCursor(QTextCursor.End)

    def keyPressEvent(self, e):

        if e.key() == Qt.Key_F1:
            self.name_rule()

        if e.key() == Qt.Key_1:
            if QApplication.keyboardModifiers() == Qt.ControlModifier:
                self.quick_distribfc_btn.setChecked(True)

        if e.key() == Qt.Key_2:
            if QApplication.keyboardModifiers() == Qt.ControlModifier:
                self.quick_distribfh_btn.setChecked(True)

        if e.key() == Qt.Key_3:
            if QApplication.keyboardModifiers() == Qt.ControlModifier:
                self.quick_distribbil_btn.setChecked(True)

        if e.key() == Qt.Key_4:
            if QApplication.keyboardModifiers() == Qt.ControlModifier:
                self.quick_distriblin_btn.setChecked(True)


        if e.key() == Qt.Key_Q:
            if QApplication.keyboardModifiers() == Qt.ControlModifier:              # test mod shortcut
                self.actionimport.trigger()
                self.check_part()
                self.pamt_dict()
                self.show_msg()
                self.pamt_GUI()
                self.append_text('进入调试模式')

        if e.key() == Qt.Key_J:
            if QApplication.keyboardModifiers() == Qt.ControlModifier:              # test mod shortcut
                self.create_tui()
                self.open_tui()

    def name_rule(self):
        reply = QMessageBox.about(self, '帮助——命名规则', '命名分为体与面的命名：\n'
                                    '请选择存在的体部件或使用快捷模板（ctrl+英文首字母），并点击导入模板\n\n'
                                    '完成后：\n'
                                    '1.对于面：选择面，并在space claim group栏下对相应的名字使用右键-replace\n'
                                    '2.对于体：请复制弹出窗口中 体名字 至space claim模型树中 指定体')

    def quick_solve(self):
        confirm_info = self.project_info_check()

        if confirm_info == True:
            self.check_part()
            self.pamt_GUI()
            self.start_btn.setText('网   格')
            self.start_btn.setEnabled(True)
            self.solver_btn.show()
            self.solver_btn.setEnabled(True)
            self.append_text('警告：进入补算模式，请先确认正确的模型或网格')

    def force_stop(self):
        try:
            if self.mesh_thread.isRunning() is True:
                self.mesh_thread.stop_mesh()
                self.clock.stop()
                self.append_text('网格划分已经被终止')

            if self.solver_thread.isRunning() is True:
                self.solver_thread.stop_solver()
                self.append_text('计算已经被终止')
        except Exception as e:
            print('not yet being running')

    def quick_distrib_judge(self):
        if self.quick_distribfc_btn.isChecked() == True:
            self.default_part_tree()

        if self.quick_distribfh_btn.isChecked() == True:
            self.default_part_tree()
            self.part_tree.topLevelItem(9).setCheckState(0, QtCore.Qt.Checked)
            self.distrib_number.setValue(2)
            self.outlet_number.setValue(1)

        if self.quick_distribbil_btn.isChecked() == True:
            self.default_part_tree()
            self.part_tree.topLevelItem(9).setCheckState(0, QtCore.Qt.Checked)
            self.distrib_number.setValue(1)
            self.outlet_number.setValue(2)

        if self.quick_distriblin_btn.isChecked() == True:
            self.default_part_tree()
            self.part_tree.topLevelItem(9).setCheckState(0, QtCore.Qt.Checked)
            self.part_tree.topLevelItem(11).setCheckState(0, QtCore.Qt.Checked)
            self.energy_checkbox.setChecked(True)
            self.outlet_number.setValue(1)
            self.distrib_number.setValue(1)

    def into_CAD(self):
        confirm_info = self.project_info_check()

        if confirm_info == True:
            self.check_part()
            self.show_msg()
            self.launchCAD()

    def project_info_check(self):
        self.pamt_dict()
        if (self.pamt['project_name'] == '') | (self.pamt['version'] == '') | (self.pamt['file_path'] == ''):
            self.append_text('请大佬将项目信息填写完全')
            return False

        if self.pamt['version'][0] != 'V':
            self.append_text('大佬，版本号首字母必须为大写V')
            return False

        if os.path.exists('%s' % (self.pamt['file_path'])) is False:
            self.append_text('项目路径不存在，请大佬检查路径信息')
            return False
        else:
            if os.path.exists('%s/project_info.py' % (self.pamt['file_path'])) is True:
                os.remove('%s/project_info.py' % (self.pamt['file_path']))

        if os.path.exists(self.pamt['cad_save_path']) == True:
            reply = QMessageBox.warning(self, "警告", "检测到路径下已存在CAD文件%s,是否覆盖？" % (self.pamt['cad_name']),
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            if reply == QMessageBox.Yes:
                result_file = 'result' + self.pamt['project_name'] + '_' + self.pamt['version']
                result_path = self.pamt['file_path'] + '\\' + result_file
                txt_out = result_path + '\\' + self.pamt['project_name'] + '.txt'
                if os.path.exists(txt_out):
                    os.remove(txt_out)
                return True
            elif reply == QMessageBox.No:
                return False

        return True

    def check_part(self):
        body_list = []
        face_list = ['inlet']
        if self.inlet_number.value() > 1:
            for i in range(self.inlet_number.value()-1):
                face_list.append('inlet%s' % (i+2))
        if self.part_tree.topLevelItem(0).checkState(0) == QtCore.Qt.Checked:
            body_list.append('inlet_sphere')
        if self.part_tree.topLevelItem(1).checkState(0) == QtCore.Qt.Checked:
            body_list.append('ai')
            if self.part_tree.topLevelItem(0).checkState(0) == QtCore.Qt.Checked:
                face_list.append('ai_in')
        if self.part_tree.topLevelItem(2).checkState(0) == QtCore.Qt.Checked:
            if self.part_tree.topLevelItem(3).checkState(0) == QtCore.Qt.Unchecked:
                print('filter and cone should be all checked')
            else:
                body_list.append('filter')
                body_list.append('cone')
                face_list.append('filter_in')
                face_list.append('filter_out')
        if self.part_tree.topLevelItem(4).checkState(0) == QtCore.Qt.Checked:
            if self.part_tree.topLevelItem(5).checkState(0) == QtCore.Qt.Unchecked:
                print('volute and fan should be all checked')
            else:
                body_list.append('volute')
                body_list.append('fan')
                if self.part_tree.topLevelItem(0).checkState(0) == QtCore.Qt.Checked:
                    face_list.append('fan_in')
                face_list.append('fan_out')
                face_list.append('fan_blade')
        if self.part_tree.topLevelItem(6).checkState(0) == QtCore.Qt.Checked:
            body_list.append('diffuser')
            if self.part_tree.topLevelItem(4).checkState(0) == QtCore.Qt.Checked:
                face_list.append('volute_out')
        if self.part_tree.topLevelItem(7).checkState(0) == QtCore.Qt.Checked:
            body_list.append('evap')
            face_list.append('evap_in')
            face_list.append('evap_out')
        if self.part_tree.topLevelItem(8).checkState(0) == QtCore.Qt.Checked:
            body_list.append('distrib')

        if self.part_tree.topLevelItem(9).checkState(0) == QtCore.Qt.Checked:
            body_list.append('hc')
            face_list.append('hc_in')
            face_list.append('hc_out')

        if self.part_tree.topLevelItem(10).checkState(0) == QtCore.Qt.Checked:
            body_list.append('ptc')
            face_list.append('ptc_in')
            face_list.append('ptc_out')
        if self.part_tree.topLevelItem(11).checkState(0) == QtCore.Qt.Checked:
            body_list.append('valve')

        face_list.append('outlet')
        if self.outlet_number.value() > 1:
            for i in range(self.outlet_number.value()-1):
                face_list.append('outlet%s'%(i+2))
        if self.distrib_number.value() > 1:
            distrib_index = body_list.index('distrib')
            body_list[distrib_index] = 'distrib1'
            for i in range(self.distrib_number.value() - 1):
                body_list.append('distrib%s' % (i + 2))

        print('body_list:%s\nface_list:%s'%(body_list, face_list))

        self.face_list = face_list
        self.body_list = body_list

    def show_msg(self):
        self.dialog_tip = Ui_tip()
        self.msg()
        self.dialog_tip.show()

    def msg(self):
        self.dialog_tip.rename_btn.clicked.connect(self.dialog_tip.close)
        self.dialog_tip.rename_btn.clicked.connect(self.update_project_info)
        self.inlet_n = self.inlet_number.value()
        self.outlet_n = self.outlet_number.value()

        self.dialog_tip.rename_table.setRowCount(max(self.inlet_n, self.outlet_n))
        self.dialog_tip.rename_table.setFixedHeight(max(self.inlet_n, self.outlet_n)*35+35)

        for i in range(self.inlet_n):
            new_item = QTableWidgetItem("%s" % (self.face_list[i]))
            self.dialog_tip.rename_table.setItem(i, 0, new_item)

        for i in range(self.outlet_n):
            new_item = QTableWidgetItem("%s" % (self.face_list[-i-1]))
            self.dialog_tip.rename_table.setItem(self.outlet_n-i-1, 1, new_item)

            new_item = QTableWidgetItem("0")
            self.dialog_tip.rename_table.setItem(self.outlet_n - i - 1, 2, new_item)

        self.dialog_tip.rename_table.customContextMenuRequested.connect(self.generate_cal_menu)

    def generate_cal_menu(self, pos):
        column_num = -1
        for i in self.dialog_tip.rename_table.selectionModel().selection().indexes():
            column_num = i.column()
            self.K_row = i.row()

        if column_num == 2:
            cal_menu = QMenu()
            cal_action = cal_menu.addAction(u"计算K值")
            action = cal_menu.exec_(self.dialog_tip.rename_table.mapToGlobal(pos))
            if action == cal_action:
                self.K_cal = Ui_cal_K()
                self.K_cal.show()
                self.K_cal.K_result.connect(self.return_K)

    def return_K(self, K):
        modify_item = QTableWidgetItem(K)
        self.dialog_tip.rename_table.setItem(self.K_row, 2, modify_item)
        self.K_cal.close()

    def update_project_info(self):
        self.outlet_list, self.K_list = [], []

        for i in range(self.inlet_n):
            self.face_list[i] = self.dialog_tip.rename_table.item(i, 0).text()

        for i in range(self.outlet_n):
            self.face_list[-i - 1] = self.dialog_tip.rename_table.item(self.outlet_n - i - 1, 1).text()
            self.outlet_list.insert(0, self.face_list[-i - 1])
            self.K_list.append(self.dialog_tip.rename_table.item(i, 2).text())

        self.K_dict = dict(zip(self.outlet_list, self.K_list))
        self.f = open('%s/project_info.py' % (self.pamt['file_path']), 'w')
        message = """
print('start script')
body_list = %s
body_number = len(body_list)

selection = Selection.Create(GetRootPart().GetAllBodies())
result = RenameObject.Execute(selection,"solid")

result = Copy.ToClipboard(Selection.Create(GetRootPart().GetAllBodies()))
result = Paste.FromClipboard()

# Delete Selection
selection = Selection.Create(GetRootPart().Components[:])
result = Delete.Execute(selection)

for i in range(body_number):
    result = Copy.ToClipboard(Selection.Create(GetRootPart().Bodies[0]))
    result = Paste.FromClipboard()

for i in range(body_number):
    selection = Selection.Create(GetRootPart().Bodies[-1-i])
    result = RenameObject.Execute(selection, body_list[i])

selection = Selection.Create(GetRootPart().Bodies[:])
result = ComponentHelper.CreateSeparateComponents(selection, None)

for i in range(body_number):
    selection = Selection.CreateByNames(body_list[i])
    result = Delete.Execute(selection)

# face rename
face_list = %s
for i in range(len(face_list)):
    primarySelection = Selection.Create(GetRootPart())
    secondarySelection = Selection()
    result = NamedSelection.Create(primarySelection, secondarySelection)

for i in range(len(face_list)):
    result = NamedSelection.Rename("Group%%s"%%(i+1), face_list[i])

# options = ShareTopologyOptions()
# options.Tolerance = MM(0.01)
# result = ShareTopology.FindAndFix(options)

# save file
options = ExportOptions.Create()
DocumentSave.Execute(r"%s", options)
print('script finished')
""" % (self.body_list, self.face_list, self.pamt['cad_save_path'])
        self.f.write(message)
        self.f.close()

    def show_c(self):
        dic = {}

        dic['evap'] = self.evap_c
        dic['hc'] = self.hc_c
        dic['valve'] = self.valve_c

        for i in self.body_list:
            try:
                dic[i].show()
            except Exception as e:
                pass

        if self.energy_checkbox.isChecked() is True:
            self.temp_c.show()

    def pamt_GUI(self):
        self.mode_info_frame.hide()
        self.return_btn.show()
        self.mass_inlet.show()
        self.show_c()

        self.start_btn.show()

    def launchCAD(self):
        self.pamt_GUI()
        self.launch_progress_display(35)
        self.append_text('正在打开CAD, 请大佬耐心等待')
        self.CAD_thread = SCDM()
        self.CAD_thread.start()
        self.CAD_thread.finishCAD.connect(self.append_text)

    def launch_progress_display(self, seconds):
        self.launch_time = timer(seconds)
        self.launch_time.start()
        self.launch_time.time_count.connect(self.time_bar)

    def time_bar(self, msg):
        self.interact_edit.undo()

        cad_progress = '正在打开CAD, 请大佬耐心等待' + ' (⊙_⊙) ' * (msg + 1)
        self.append_text(cad_progress)

    def import_info(self):
        path = QFileDialog.getOpenFileName(self, '选择要输入的Excel模板',
                                           r'C:\Users\BZMBN4\Desktop', 'Excel Files (*.xlsx; *.xls; *.xlsm)')
        if path[0] != '':
            excel_path = path[0]
            info, squence = self.excel_import(excel_path)

            for i in info:
                name = i + '_edit'
                widget = self.findChild(QLineEdit, name)
                if widget != None:
                    widget.setText(str(info[i]))

            self.project_name_edit.setText(info['project_name'])
            self.version_name_edit.setText('V')

            self.append_text('Excel:%s导入成功' % path[0])

    def excel_import(self, excel_path):
        import xlrd
        wb = xlrd.open_workbook(excel_path)
        sheet = wb.sheet_by_name('info')

        nrows = sheet.nrows
        row_squence = list(range(1, nrows + 1))

        info = dict(zip(sheet.col_values(0), sheet.col_values(1)))
        squence = dict(zip(sheet.col_values(0), row_squence))

        return info, squence

    def case_address(self):
        case_out = QFileDialog.getExistingDirectory(self, '选择项目路径', 'C:/Users/BZMBN4/Desktop/')
        self.case_path = case_out
        self.project_address_edit.setText(self.case_path)

    def export_pamt(self):
        self.pamt_dict()

        path = QFileDialog.getSaveFileName(self, directory='%s'%(self.project_address_edit.text()),
                                           filter='Excel, *.xlsx')
        try:
            excel_save_path = path[0]
            print(excel_save_path)
            self.create_excel(excel_save_path)
            os.system(excel_save_path)
        except Exception as e:
            self.append_text('导出地址有错误，请重新选择')

    def create_excel(self, excel_name):
        import openpyxl
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet(title='info', index=0)
        excel_info = list(self.pamt.keys())
        excel_pamt = list(self.pamt.values())
        for i in range(len(excel_info)):
            worksheet.cell(i+1, 1, excel_info[i])
            worksheet.cell(i+1, 2, excel_pamt[i])

        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 40

        workbook.save(filename=excel_name)
        self.append_text('已创建新Excel:%s 新表名: info' % (excel_name))
        self.interact_edit.moveCursor(QTextCursor.End)

    def unit_convert(self):
        self.volume_unit = Ui_unit()
        self.volume_unit.show()
        self.volume_unit.unit_convert_result.connect(self.volume_input)

    def volume_input(self, msg):
        self.mass_inlet_edit.setText(msg)

    def pamt_dict(self):
        self.pamt['project_name'] = self.project_name_edit.text()
        self.pamt['version'] = self.version_name_edit.text()
        self.pamt['file_path'] = self.project_address_edit.text()
        self.pamt['edit_time'] = self.version_date_edit.text()
        self.pamt['cad_name'] = self.pamt['project_name'] + '_' + self.pamt['version'] + '_' + self.pamt['edit_time']
        self.pamt['cad_save_path'] = self.pamt['file_path'] + '/' + self.pamt['cad_name'] + '.scdoc'
        self.pamt['mass_inlet'] = self.mass_inlet_edit.text()

        if 'evap' in self.body_list:
            self.pamt['evap_c1'] = self.evap_c1_edit.text()
            self.pamt['evap_c2'] = self.evap_c2_edit.text()
            self.pamt['evap_x1'] = self.evap_x1_edit.text()
            self.pamt['evap_y1'] = self.evap_y1_edit.text()
            self.pamt['evap_z1'] = self.evap_z1_edit.text()
            self.pamt['evap_x2'] = self.evap_x2_edit.text()
            self.pamt['evap_y2'] = self.evap_y2_edit.text()
            self.pamt['evap_z2'] = self.evap_z2_edit.text()

        if 'hc' in self.body_list:
            self.pamt['hc_c1'] = self.hc_c1_edit.text()
            self.pamt['hc_c2'] = self.hc_c2_edit.text()
            self.pamt['hc_x1'] = self.hc_x1_edit.text()
            self.pamt['hc_y1'] = self.hc_y1_edit.text()
            self.pamt['hc_z1'] = self.hc_z1_edit.text()
            self.pamt['hc_x2'] = self.hc_x2_edit.text()
            self.pamt['hc_y2'] = self.hc_y2_edit.text()
            self.pamt['hc_z2'] = self.hc_z2_edit.text()

        if 'valve' in self.body_list:
            self.pamt['valve_ox'] = self.valve_ox_edit.text()
            self.pamt['valve_oy'] = self.valve_oy_edit.text()
            self.pamt['valve_oz'] = self.valve_oz_edit.text()
            self.pamt['valve_dx'] = self.valve_dx_edit.text()
            self.pamt['valve_dy'] = self.valve_dy_edit.text()
            self.pamt['valve_dz'] = self.valve_dz_edit.text()
            self.pamt['valve_td'] = self.valve_td_edit.text()
            self.pamt['valve_rp'] = self.valve_rp_edit.text()
            self.pamt['valve_sa'] = self.valve_sa_edit.text()
            self.pamt['valve_fa'] = self.valve_fa_edit.text()

        if self.energy_checkbox.isChecked() is True:
            self.pamt['temp_inlet'] = self.temp_inlet_edit.text()
            self.pamt['temp_hc'] = self.temp_hc_edit.text()

    def create_tui(self):
        self.pamt_dict()
        d = self.pamt
        self.check_part()

        whole_jou = ''
        project_title = d['project_name']
        version_name = d['version']
        cad_name = d['cad_name']

        case_out = d['file_path']

        self.jou_mesh_path = d['file_path'] + '/' + project_title + '-' + version_name + '-mesh-TUI.jou'  # txt final path
        print('output journal in:', d['file_path'])
        jou_mesh = open(self.jou_mesh_path, 'w')

        if 'hc' in self.body_list:
            internal_list = ['evap*', 'hc*']

            uni_face_list = ['evap_in', 'evap_out', 'hc_out']
            pressure_face_list = ['*inlet*', 'evap_in', 'evap_out', 'hc_in', 'hc_out', '*outlet*']

        else:
            internal_list = ['evap*']
            uni_face_list = ['evap_in', 'evap_out']
            pressure_face_list = ['*inlet*', 'evap_in', 'evap_out', '*outlet*']
        dead_zone_list = []
        if 'valve' in self.body_list:
            dead_zone_list.append('valve')
            self.body_list.remove('valve')
        mesh_zone_list = self.body_list

        CFD = fluent_tui.tui(whole_jou, project_title, version_name, case_out, cad_name)
        CFD.mesh.import_distrib()
        CFD.mesh.general_improve()
        CFD.mesh.fix_slivers()
        CFD.mesh.general_improve()
        CFD.mesh.compute_volume_region()
        CFD.mesh.volume_mesh_change_type(dead_zone_list)
        if self.energy_checkbox.isChecked() is True:
            CFD.mesh.retype_face(face_list=['hc*'], face_type='radiator')
            internal_list.remove('hc*')
            CFD.mesh.auto_mesh_volume(1.25, 'poly')
        else:
            CFD.mesh.auto_mesh_volume()
        CFD.mesh.auto_node_move()
        CFD.mesh.rename_cell(zone_list=mesh_zone_list)
        CFD.mesh.retype_face(face_list=['inlet*'], face_type='pressure-inlet')
        CFD.mesh.retype_face(face_list=internal_list, face_type='internal')
        CFD.mesh.retype_face(face_list=['outlet*'], face_type='outlet-vent')
        CFD.mesh.check_quality()
        CFD.mesh.prepare_for_solve()
        CFD.mesh.write_mesh()
        CFD.close_fluent()

        jou_mesh.write(CFD.whole_jou)
        jou_mesh.close()

        self.jou_solve_path = d['file_path'] + '/' + project_title + '-' + version_name + '-solve-TUI.jou'
        print('output journal in:', d['file_path'])
        jou_solve = open(self.jou_solve_path, 'w')

        evap_d1 = [d['evap_x1'], d['evap_y1'], d['evap_z1']]
        evap_d2 = [d['evap_x2'], d['evap_y2'], d['evap_z2']]
        mass_flux_list = ['inlet*', 'outlet*']

        CFD = fluent_tui.tui(whole_jou, project_title, version_name, case_out, cad_name)
        CFD.setup.read_mesh()
        CFD.setup.rescale()
        CFD.setup.turb_models()
        CFD.setup.porous_zone('evap', evap_d1, evap_d2, d['evap_c1'], d['evap_c2'])
        if 'hc' in self.body_list:
            hc_d1 = [d['hc_x1'], d['hc_y1'], d['hc_z1']]
            hc_d2 = [d['hc_x2'], d['hc_y2'], d['hc_z2']]
            CFD.setup.porous_zone('hc', hc_d1, hc_d2, d['hc_c1'], d['hc_c2'])
        CFD.setup.BC_type('inlet', 'mass-flow-inlet')
        CFD.setup.BC_type('outlet*()', 'outlet-vent')
        CFD.setup.BC_mass_flow_inlet('inlet', d['mass_inlet'])
        for i in self.K_dict:
            CFD.setup.BC_outlet_vent(self.K_dict[i], i)
        CFD.setup.solution_method()
        if self.energy_checkbox.isChecked() is True:
            inlet_temp = float(d['temp_inlet']) + 273.15
            hc_temp = float(d['temp_hc']) + 273.15
            CFD.setup.energy_eqt('yes')
            CFD.setup.init_temperature('mass-flow-inlet', 'outlet-vent', inlet_temp)
            CFD.setup.heat_flux('hc_in', hc_temp)
            CFD.setup.heat_flux('hc_out', hc_temp)
            CFD.setup.report_definition('temperature', 'surface-areaavg', ['outlet*'], 'yes', 'temperature')
        CFD.setup.report_definition('volume', 'surface-volumeflowrate', ['inlet*'])
        CFD.setup.report_definition('mass-flux', 'surface-massflowrate', mass_flux_list, 'no')
        CFD.setup.report_definition('pressure', 'surface-areaavg', ['evap_in'])
        CFD.setup.convergence_criterion()
        CFD.setup.hyb_initialize()
        CFD.setup.start_calculate(230)
        CFD.setup.write_case_data()

        volume_face_list = ['inlet*', 'outlet*']

        CFD.post.create_result_file()
        CFD.post.set_background()
        if self.energy_checkbox.isChecked() is True:
            CFD.post.txt_surface_integrals('area-weighted-avg', ['outlet*'], 'temperature')
            CFD.post.create_streamline('temp_pathline', 'inlet', '', 'temperature')
            CFD.post.snip_avz(8, 'temp_pathline')
        else:
            CFD.post.create_contour('evap_out', 'evap_out')
            if 'hc' in self.body_list:
                CFD.post.create_contour('hc_out', 'hc_out')

            CFD.post.create_streamline('whole_pathline', 'inlet')
            CFD.post.create_streamline('distrib_pathline', 'evap_out', [0, 15])
            CFD.post.snip_avz(5, 'whole_pathline')
            CFD.post.snip_avz(6, 'distrib_pathline')
            CFD.post.snip_avz(7, 'evap_out')
            if 'hc' in self.body_list:
                CFD.post.snip_avz(9, 'hc_out')
            CFD.post.snip_model(10, 'model')
        CFD.post.txt_surface_integrals('volume-flow-rate', volume_face_list)
        CFD.post.txt_mass_flux()
        CFD.post.txt_surface_integrals('uniformity-index-area-weighted', uni_face_list, 'velocity-magnitude')
        CFD.post.txt_surface_integrals('area-weighted-avg', pressure_face_list, 'total-pressure')
        CFD.post.txt_surface_integrals('area-weighted-avg', pressure_face_list, 'pressure')
        CFD.post.snip_mode_off()
        CFD.close_fluent()

        jou_solve.write(CFD.whole_jou)
        jou_solve.close()

    def open_tui(self):
        os.system(self.jou_mesh_path)
        os.system(self.jou_solve_path)

    def begin(self):
        self.start_btn.setDisabled(True)
        self.create_tui()

        self.mesh_condition = '启动fluent'
        self.append_text(self.mesh_condition)

        self.mesh_thread = fluent_mesh(self.jou_mesh_path)
        self.mesh_thread.start()
        self.mesh_clock()
        self.mesh_thread.mesh_feedback.connect(self.mesh_msg)
        self.mesh_thread.mesh_timeuse.connect(self.mesh_time_use)
        self.mesh_thread.mesh_finish.connect(self.mesh_finish_msg)

        self.actionstop.setEnabled(True)

    def mesh_clock(self):
        try:
            size = self.get_FileSize(self.pamt['cad_save_path'])
            print("文件路径：%s\n大小：%s MB" % (self.pamt['cad_save_path'], size))
        except Exception as e:
            self.append_text('模型文件未找到，请检查')
        mesh_time = 120
        self.timer(mesh_time)
        self.time_running.connect(self.mesh_clock_show)

    def get_FileSize(self, file_path):
        fsize = os.path.getsize(file_path)
        fsize = fsize / float(1024 * 1024)
        return round(fsize, 3)

    def mesh_clock_show(self, msg):
        dot_list = ['.   ', '..  ', '... ', '....']
        remain_time = self.total_time - msg
        dot_circle = dot_list[msg % 4]
        remain_time_string = '   网格阶段剩余%s秒' % (remain_time)
        clock_msg = self.mesh_condition + dot_circle + remain_time_string

        self.interact_edit.undo()
        self.append_text(clock_msg)

    def mesh_msg(self, msg):
        self.mesh_condition = msg

    def mesh_time_use(self, msg):
        self.step = msg

    def mesh_finish_msg(self, msg):
        self.interact_edit.undo()
        self.append_text(msg)
        self.solver_btn.click()

    def solver(self):
        self.start_btn.setDisabled(True)
        self.solver_btn.setDisabled(True)
        self.actionstop.setEnabled(True)

        self.solver_condition = '启动fluent...'
        self.append_text(self.solver_condition)

        self.solver_thread = fluent_solver(self.jou_solve_path)
        self.solver_thread.start()
        # self.solver_clock()
        # self.solver_thread.solver_feedback.connect(self.solver_msg)
        # self.solver_thread.solver_timeuse.connect(self.solver_time_use)
        # self.solver_thread.solver_finish.connect(self.solver_finish_msg)

    def timer(self, total_seconds):
        self.clock = QBasicTimer()
        self.step = 0
        self.clock.start(1000, self)
        self.total_time = total_seconds

    def timerEvent(self, e):
        if self.step >= self.total_time:
            self.clock.stop()
            return

        self.step = self.step + 1
        self.time_running.emit(self.step)


class Ui_tip(Ui_tip_widget, QWidget):
    def __init__(self):
        super(Ui_tip_widget, self).__init__()
        self.setupUi(self)


class Ui_cal_K(Ui_K_calculator, QWidget):
    K_result = pyqtSignal(str)

    def __init__(self):
        super(Ui_K_calculator, self).__init__()
        self.setupUi(self)
        self.cal_method()
        self.QP_btn.click()

    def initialize(self):
        self.R_frame.hide()
        self.QP_frame.hide()

    def cal_method(self):
        self.R_btn.clicked.connect(lambda: self.choosen_method(self.R_frame, self.R_method))
        self.QP_btn.clicked.connect(lambda: self.choosen_method(self.QP_frame, self.QP_method))

    def choosen_method(self, show_frame, method):
        self.initialize()
        show_frame.show()
        self.K_cal_btn.disconnect()
        self.K_cal_btn.clicked.connect(method)

    def QP_method(self):
        ls = float(self.volume_edit.text())
        mm2 = float(self.area_edit.text())
        p = float(self.pressure_edit.text())
        rho = 1.225

        m3s = ls / 1000
        m2 = mm2/1000/1000
        v = m3s/m2

        K = 2*p/rho/v**2
        self.K_result.emit("%.3f" % K)

    def R_method(self):
        r = float(self.R_edit.text())
        mm2 = float(self.area_edit.text())
        rho = 1.225

        m2 = mm2 / 1000 / 1000

        K = 1000*2*r*m2**2/rho
        self.K_result.emit("%.3f" % K)


class Ui_unit(Ui_unit_converter, QWidget):
    unit_convert_result = pyqtSignal(str)

    def __init__(self):
        super(Ui_unit, self).__init__()
        self.setupUi(self)
        self.unit_btn()
        self.default_state()

    def default_state(self):
        self.l_btn.click()
        self.s_btn.click()

    def unit_btn(self):
        self.kg_btn.clicked.connect(lambda: self.volume_label_show('Kg', 1))
        self.m3_btn.clicked.connect(lambda: self.volume_label_show('M3', 1.225))
        self.l_btn.clicked.connect(lambda: self.volume_label_show('L', 0.001225))

        self.h_btn.clicked.connect(lambda: self.time_label_show('H', 3600))
        self.min_btn.clicked.connect(lambda: self.time_label_show('Min', 60))
        self.s_btn.clicked.connect(lambda: self.time_label_show('S', 1))

        self.confirm_btn.clicked.connect(self.calculate)

    def volume_label_show(self, text, factor):
        self.volume_label.setText(text)
        self.volume_factor = factor

    def time_label_show(self, text, factor):
        self.time_label.setText(text)
        self.time_factor = factor

    def calculate(self):
        if self.value_edit.text() == '':
            value = 0
        else:
            value = float(self.value_edit.text())
        result = round(value*self.volume_factor/self.time_factor, 5)
        self.unit_convert_result.emit(str(result))


class Ui_porous(Ui_porous_model_form, QWidget):
    def __init__(self):
        super(Ui_porous_model_form, self).__init__()
        self.setupUi(self)
        self.default_ui()
        self.default_btn()

    def default_ui(self):
        self.operate_frame.hide()
        self.resize(220, 135)

    def default_btn(self):
        self.model_combox.activated.connect(self.choose)
        self.QP_table.customContextMenuRequested.connect(self.generate_unit_menu)
        self.cal_btn.clicked.connect(self.cal_C1C2)
        self.unit_choose = 'kg/h'

    def choose(self, i):
        if i == 0:
            self.add_mode()

    def add_mode(self):
        self.operate_frame.show()
        self.resize(670, 670)
        self.del_btn.hide()

        self.modify_btn.setText("添加")
        self.modify_btn.clicked.connect(self.modify_porous)

    def generate_unit_menu(self, pos):
        column_num = -1
        for i in self.QP_table.selectionModel().selection().indexes():
            column_num = i.column()

        if column_num == 0:
            unit_menu = QMenu()
            unit_ls = unit_menu.addAction(u"l/s")
            unit_mh = unit_menu.addAction(u"m3/h")
            unit_kgm = unit_menu.addAction(u"kg/min")
            unit_kgh = unit_menu.addAction(u"kg/h")
            action = unit_menu.exec_(self.QP_table.mapToGlobal(pos))
            try:
                self.unit_choose = action.text()
                unit_item = QTableWidgetItem("流量(%s)" % action.text())
                self.QP_table.setHorizontalHeaderItem(0, unit_item)
            except:
                pass

    def cal_C1C2(self):
        l_raw = self.length_edit.text()
        w_raw = self.width_edit.text()
        h_raw = self.height_edit.text()
        rho = 1.225                                 # density
        mu = 0.000017894                            # dynamic viscosity
        self.Q_unit = {'l/s': 1000, 'm3/h': 1 / 3600, 'kg/min': 1 / rho / 60, 'kg/h': 1 / rho / 3600}

        if l_raw and w_raw and h_raw != '':
            l = float(l_raw)/1000
            w = float(w_raw)/1000
            h = float(h_raw)/1000
            eff_area = l*w                          # effective area size

            import numpy as np
            row_num = self.QP_table.rowCount()
            Q = np.zeros(row_num + 1)
            P = np.zeros(row_num + 1)
            for i in range(row_num):
                q = self.QP_table.item(i, 0).text()
                p = self.QP_table.item(i, 1).text()
                try:
                    q = float(q)
                    Q[i + 1] = q
                except:
                    continue
                try:
                    p = float(p)
                    P[i + 1] = p
                except:
                    continue

            Q = set(Q)
            Q = np.array(list(Q))
            Q.sort()

            P = set(P)
            P = np.array(list(P))
            P.sort()
            v = Q*self.Q_unit[self.unit_choose]/eff_area

            print('v:', v)
            print('P:', P)
            self.plot_frame.setVisible(True)
            self.plot_frame.mpl.start_static_plot(v, P)
            a = self.plot_frame.mpl.a
            b = self.plot_frame.mpl.b
            C1 = b/h/mu
            C2 = 2*a/rho/h
            self.c1_edit.setText(str('%.2e'%C1))
            self.c2_edit.setText(str('%.2f'%C2))
        else:
            print('please complete all blank edit')

    def modify_porous(self):
        model_name = self.model_name_edit.text()
        self.model_combox.addItem(model_name)


class SCDM(QThread):
    finishCAD = pyqtSignal(str)

    def __init__(self, parent=None):
        super(SCDM, self).__init__(parent)

    def run(self):
        import subprocess
        p = subprocess.Popen(r'C:\Program Files\ANSYS Inc\v191\scdm\SpaceClaim.exe', shell=True,
                             stdout=subprocess.PIPE)
        out, err = p.communicate()
        out = out.decode()
        self.finishCAD.emit('CAD软件已关闭')


class timer(QThread):
    time_count = pyqtSignal(int)

    def __init__(self, second):
        super(timer, self).__init__()
        self.second = second
        self.timer = QBasicTimer()
        self.step = 0
        self.timer.start(1000, self)

    def run(self):
        pass

    def timerEvent(self, e):
        if self.step >= self.second:
            self.timer.stop()
            return

        self.step = self.step + 1
        self.time_count.emit(self.step)


class fluent_mesh(QThread):
    mesh_feedback = pyqtSignal(str)
    mesh_timeuse = pyqtSignal(int)
    mesh_finish = pyqtSignal(str)

    def __init__(self, tui):
        super(fluent_mesh, self).__init__()
        self.tui = tui

    def run(self):
        start_time = time.strftime('%M:%S', time.localtime(time.time()))
        print(start_time)
        import subprocess
        self.p = subprocess.Popen(r'cd C:\\Program Files\\ANSYS Inc\\v191\\fluent\\ntbin\\win64 && '
                           r'fluent 3d -meshing -t4 -gu -i %s' % (self.tui),
                                  shell=True, stdout=subprocess.PIPE)

        nl = 0
        finish_count = 0
        while self.p.poll() == None:
            nl += 1
            line = self.p.stdout.readline()
            self.msg = line.decode()
            print(nl, self.msg)

            if (nl > 30) & (nl <= 50):
                self.stage_report('Cleanup script file is', 'load_time_begin', '载入模型', 29)
            elif (nl > 50) & (nl <= 130):
                self.stage_report('They will be imported in whatever units they were created', 'face_mesh_time_start',
                '处理面网格', 80)
            elif (nl > 130) & (nl <= 200):
                if 'the previous max quality' in self.msg:
                    msg_strip = self.msg.strip()
                    face_mesh_quality = msg_strip[-8:]
                self.stage_report('/objects/volumetric-regions/compute', 'volume_mesh_time_start',
                                  '处理体网格', 83)
            elif (nl > 200) & (nl <= 420):
                self.stage_report('/mesh/check-quality', 'mesh_finish_time', '网格生成完毕', 113)
                if 'Minimum Orthogonal Quality' in self.msg:
                    msg_strip = self.msg.strip()
                    volume_mesh_quality = float(msg_strip[-11:-7])/10.0
            elif (nl > 420) & (finish_count < 2):
                if 'Writing "' in self.msg:
                    finish_count = 1
                if finish_count == 1:
                    if 'Done.' in self.msg:
                        finish_count = 2
                        end_time = time.strftime('%M:%S', time.localtime(time.time()))
                        print(end_time)
                        print('总共有%s行输出语句' % nl)
        else:
            self.mesh_timeuse.emit(120)
            self.mesh_finish.emit('网格生成完毕')

    def stage_report(self, stage_marker, stage_name, report_msg, timeuse):
        if stage_marker in self.msg:
            stage_time = time.strftime('%M:%S', time.localtime(time.time()))
            self.mesh_feedback.emit(report_msg)
            self.mesh_timeuse.emit(timeuse)
            print('%s%s' % (stage_name, stage_time))

    def stop_mesh(self):
        self.p.terminate()
    

class fluent_solver(QThread):
    solver_feedback = pyqtSignal(str)

    def __init__(self, tui):
        super(fluent_solver, self).__init__()
        self.tui = tui

    def run(self):

        import subprocess
        self.p = subprocess.Popen(r'cd C:\\Program Files\\ANSYS Inc\\v191\\fluent\\ntbin\\win64 && '
                           r'fluent 3d -t12 -gu -i %s' % (self.tui), shell=True, stdout=subprocess.PIPE)

        nl = 0
        finish_count = 0
        while self.p.poll() == None:
            line = self.p.stdout.readline()
            self.msg = line.decode()
            print(nl, self.msg)
            nl += 1

        print('总共有%s行输出语句'%nl)

    def stop_solver(self):
        self.p.terminate()


if __name__ == "__main__":
    try:
        app = QApplication(sys.argv)
        myWin = MyMainWindow()
        myWin.show()
        sys.exit(app.exec_())
    except:
        import traceback
        traceback.print_exc()

